import csv
import io
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from typing import Optional
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from ..database import get_db
from ..models.animal import Animal, StatusEnum, SexoEnum
from ..models.pesagem import Pesagem
from ..models.saude import Saude
from ..models.movimentacao import Movimentacao, TipoMovEnum
from ..models.despesa_fixa import DespesaFixa
from ..models.custo_nutricional import CustoNutricional
from ..models.lote import Lote
from ..auth import get_current_user, check_assinatura_ativa
from ..models.user import User
from .pesagens import _calcular_gmd
from calendar import monthrange

router = APIRouter()


# ── Helpers ──────────────────────────────────────────────────────────────────

def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _pdf_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


HEADER_FILL = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _format_header(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_w: int = 40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)


def _overlap_days(rec_inicio: date, rec_fim: Optional[date], per_inicio: date, per_fim: date) -> int:
    inicio = max(rec_inicio, per_inicio)
    fim = min(rec_fim or per_fim, per_fim)
    return max(0, (fim - inicio).days + 1) if fim >= inicio else 0


@router.get("/animais.csv")
def exportar_animais(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    animais = (
        db.query(Animal)
        .filter(Animal.user_id == current_user.id, Animal.deletado_em == None)
        .order_by(Animal.brinco)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Brinco", "Nome", "Raça", "Sexo", "Status", "Lote ID",
        "Data Nascimento", "Peso Entrada (kg)", "Origem", "Observações", "Cadastrado em"
    ])
    for a in animais:
        writer.writerow([
            a.brinco, a.nome or "", a.raca or "", a.sexo, a.status,
            a.lote_id or "", a.data_nascimento or "", a.peso_entrada or "",
            a.origem or "", a.observacoes or "",
            a.created_at.strftime("%d/%m/%Y") if a.created_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=animais_{date.today()}.csv"},
    )


@router.get("/pesagens.csv")
def exportar_pesagens(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    pesagens = (
        db.query(Pesagem)
        .join(Animal)
        .filter(Animal.user_id == current_user.id)
        .order_by(Animal.brinco, Pesagem.data)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Brinco Animal", "Data", "Peso (kg)", "GMD", "Observações"])
    for p in pesagens:
        gmd = _calcular_gmd(db, p.animal_id, p, current_user.id)
        writer.writerow([
            p.animal.brinco if p.animal else p.animal_id,
            p.data, p.peso_kg, gmd if gmd is not None else "", p.observacoes or "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=pesagens_{date.today()}.csv"},
    )


@router.get("/financeiro.csv")
def exportar_financeiro(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    movs = (
        db.query(Movimentacao)
        .join(Animal)
        .filter(Animal.user_id == current_user.id)
        .order_by(Movimentacao.data.desc())
        .all()
    )

    saudes = (
        db.query(Saude)
        .join(Animal)
        .filter(Animal.user_id == current_user.id, Saude.custo != None)
        .order_by(Saude.data.desc())
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Data", "Tipo", "Animal (Brinco)", "Valor (R$)", "Peso (kg)", "Descrição"])

    for m in movs:
        writer.writerow([
            m.data, m.tipo,
            m.animal.brinco if m.animal else m.animal_id,
            m.valor or "", m.peso_kg or "",
            f"{m.origem or ''} → {m.destino or ''}".strip(" →"),
        ])

    for s in saudes:
        writer.writerow([
            s.data, f"saude/{s.tipo}",
            s.animal.brinco if s.animal else s.animal_id,
            s.custo or "", "", s.descricao,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=financeiro_{date.today()}.csv"},
    )


# ── Excel exports ────────────────────────────────────────────────────────────

@router.get("/animais.xlsx")
def exportar_animais_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    animais = (
        db.query(Animal)
        .filter(Animal.user_id == current_user.id, Animal.deletado_em == None)
        .order_by(Animal.brinco)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Animais"
    ws.append([
        "Brinco", "Nome", "Raça", "Sexo", "Categoria", "Status", "Lote ID",
        "Data Nascimento", "Peso Entrada (kg)", "Origem", "Observações", "Cadastrado em",
    ])
    for a in animais:
        ws.append([
            a.brinco, a.nome or "", a.raca or "", a.sexo.value if hasattr(a.sexo, "value") else str(a.sexo),
            a.categoria.value if a.categoria else "",
            a.status.value if hasattr(a.status, "value") else str(a.status),
            a.lote_id or "",
            a.data_nascimento.isoformat() if a.data_nascimento else "",
            a.peso_entrada or "",
            a.origem or "", a.observacoes or "",
            a.created_at.strftime("%d/%m/%Y") if a.created_at else "",
        ])
    _format_header(ws)
    _autosize(ws)
    return _xlsx_response(wb, f"animais_{date.today()}.xlsx")


@router.get("/pesagens.xlsx")
def exportar_pesagens_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    pesagens = (
        db.query(Pesagem).join(Animal)
        .filter(Animal.user_id == current_user.id)
        .order_by(Animal.brinco, Pesagem.data)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Pesagens"
    ws.append(["Brinco", "Data", "Peso (kg)", "GMD (kg/dia)", "Observações"])
    for p in pesagens:
        gmd = _calcular_gmd(db, p.animal_id, p, current_user.id)
        ws.append([
            p.animal.brinco if p.animal else str(p.animal_id),
            p.data.isoformat() if p.data else "",
            p.peso_kg, gmd if gmd is not None else "", p.observacoes or "",
        ])
    _format_header(ws)
    _autosize(ws)
    return _xlsx_response(wb, f"pesagens_{date.today()}.xlsx")


@router.get("/financeiro.xlsx")
def exportar_financeiro_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    movs = (
        db.query(Movimentacao).join(Animal)
        .filter(Animal.user_id == current_user.id)
        .order_by(Movimentacao.data.desc())
        .all()
    )
    saudes = (
        db.query(Saude).join(Animal)
        .filter(Animal.user_id == current_user.id, Saude.custo != None)
        .order_by(Saude.data.desc())
        .all()
    )

    wb = Workbook()
    ws_mov = wb.active
    ws_mov.title = "Movimentações"
    ws_mov.append([
        "Data", "Tipo", "Brinco", "Valor (R$)", "Frete (R$)", "Desconto (R$)",
        "Peso (kg)", "Preço @ (R$)", "Origem", "Destino", "Observações",
    ])
    for m in movs:
        ws_mov.append([
            m.data.isoformat() if m.data else "",
            m.tipo.value if hasattr(m.tipo, "value") else str(m.tipo),
            m.animal.brinco if m.animal else str(m.animal_id),
            m.valor or "", m.frete or "", m.desconto or "",
            m.peso_kg or "", m.preco_arroba or "",
            m.origem or "", m.destino or "", m.observacoes or "",
        ])
    _format_header(ws_mov)
    _autosize(ws_mov)

    ws_saude = wb.create_sheet("Custos de Saúde")
    ws_saude.append(["Data", "Brinco", "Tipo", "Descrição", "Custo (R$)"])
    for s in saudes:
        ws_saude.append([
            s.data.isoformat() if s.data else "",
            s.animal.brinco if s.animal else str(s.animal_id),
            s.tipo.value if hasattr(s.tipo, "value") else str(s.tipo),
            s.descricao, s.custo or "",
        ])
    _format_header(ws_saude)
    _autosize(ws_saude)

    return _xlsx_response(wb, f"financeiro_{date.today()}.xlsx")


# ── PDF exports ──────────────────────────────────────────────────────────────

def _styles():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", parent=base["Title"], fontSize=18, textColor=colors.HexColor("#2D6A4F")),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontSize=12, textColor=colors.HexColor("#1B4332"), spaceAfter=6),
        "body": ParagraphStyle("body", parent=base["BodyText"], fontSize=9.5),
        "muted": ParagraphStyle("muted", parent=base["BodyText"], fontSize=8.5, textColor=colors.HexColor("#6b7280")),
    }


def _table_style(header_bg: str = "#2D6A4F"):
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_bg)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDF4")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8F3DC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])


@router.get("/animais.pdf")
def exportar_animais_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    animais = (
        db.query(Animal)
        .filter(
            Animal.user_id == current_user.id,
            Animal.deletado_em == None,  # noqa: E711
            Animal.status == StatusEnum.ativo,
        )
        .order_by(Animal.brinco)
        .all()
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        title="Rebanho", author="BovIA",
    )
    st = _styles()
    elems = [
        Paragraph(f"Rebanho — {current_user.fazenda_nome or ''}", st["title"]),
        Paragraph(f"{len(animais)} animal(is) ativos · gerado em {date.today().strftime('%d/%m/%Y')}", st["muted"]),
        Spacer(1, 0.4 * cm),
    ]

    data_rows = [["Brinco", "Nome", "Raça", "Sexo", "Categoria", "Status", "Peso (kg)", "Nasc.", "Origem"]]
    for a in animais:
        data_rows.append([
            a.brinco or "—",
            a.nome or "—",
            a.raca or "—",
            "Macho" if (a.sexo and a.sexo.value if hasattr(a.sexo, "value") else str(a.sexo)) == "macho" else "Fêmea",
            (a.categoria.value.replace("_", " ").title() if a.categoria else "—"),
            (a.status.value if hasattr(a.status, "value") else str(a.status)).title(),
            f"{a.peso_entrada:.0f}" if a.peso_entrada else "—",
            a.data_nascimento.strftime("%d/%m/%y") if a.data_nascimento else "—",
            a.origem or "—",
        ])
    table = Table(data_rows, repeatRows=1, colWidths=[2.2 * cm, 4 * cm, 3 * cm, 1.8 * cm, 2.6 * cm, 2 * cm, 1.8 * cm, 2 * cm, 3 * cm])
    table.setStyle(_table_style())
    elems.append(table)

    doc.build(elems)
    return _pdf_response(buf, f"rebanho_{date.today()}.pdf")


@router.get("/resumo-contador.pdf")
def resumo_contador_pdf(
    data_inicio: date = Query(..., description="Início do período (YYYY-MM-DD)"),
    data_fim: date = Query(..., description="Fim do período (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if data_fim < data_inicio:
        raise HTTPException(400, "data_fim anterior a data_inicio")

    uid = current_user.id

    # Movimentações no período
    movs = (
        db.query(Movimentacao).join(Animal)
        .filter(
            Animal.user_id == uid,
            Movimentacao.data >= data_inicio,
            Movimentacao.data <= data_fim,
        )
        .order_by(Movimentacao.data)
        .all()
    )
    vendas = [m for m in movs if (m.tipo.value if hasattr(m.tipo, "value") else str(m.tipo)) == "venda"]
    compras = [m for m in movs if (m.tipo.value if hasattr(m.tipo, "value") else str(m.tipo)) == "compra"]

    receita_bruta = sum(v.valor or 0 for v in vendas)
    descontos = sum(v.desconto or 0 for v in vendas)
    receita_liquida = receita_bruta - descontos
    custo_compras = sum(c.valor or 0 for c in compras)
    fretes = sum(c.frete or 0 for c in compras)

    # Custos de saúde
    saudes = (
        db.query(Saude).join(Animal)
        .filter(
            Animal.user_id == uid, Saude.custo != None,
            Saude.data >= data_inicio, Saude.data <= data_fim,
        )
        .all()
    )
    custo_saude = sum(s.custo or 0 for s in saudes)

    # Despesas fixas pro-rata por dias do periodo
    dias_periodo = (data_fim - data_inicio).days + 1
    despesas = db.query(DespesaFixa).filter(DespesaFixa.user_id == uid).all()
    custo_despesas_op = 0.0
    custo_impostos = 0.0
    for d in despesas:
        dias = _overlap_days(d.data_inicio, d.data_fim, data_inicio, data_fim)
        if dias <= 0:
            continue
        valor_diario = (d.valor_mensal or 0) / 30.0
        rateio = valor_diario * dias
        cat = d.categoria.value if hasattr(d.categoria, "value") else str(d.categoria)
        if cat == "impostos":
            custo_impostos += rateio
        else:
            custo_despesas_op += rateio

    custo_total = custo_compras + fretes + custo_saude + custo_despesas_op
    lucro_bruto = receita_liquida - custo_total
    lucro_liquido = lucro_bruto - custo_impostos

    # PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title="Resumo Contábil", author="BovIA",
    )
    st = _styles()
    elems = [
        Paragraph("Resumo Contábil", st["title"]),
        Paragraph(
            f"{current_user.fazenda_nome or ''} · período {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} ({dias_periodo} dias)",
            st["muted"],
        ),
        Spacer(1, 0.5 * cm),
    ]

    def _money(v: float) -> str:
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    # Receita
    elems.append(Paragraph("Receita", st["h2"]))
    receita_table = Table([
        ["Vendas brutas", _money(receita_bruta)],
        ["(–) Descontos concedidos", _money(descontos)],
        ["Receita líquida", _money(receita_liquida)],
    ], colWidths=[12 * cm, 5 * cm])
    receita_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.HexColor("#2D6A4F")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elems.append(receita_table)
    elems.append(Spacer(1, 0.4 * cm))

    # Custos
    elems.append(Paragraph("Custos operacionais", st["h2"]))
    custos_table = Table([
        ["Compras de animais", _money(custo_compras)],
        ["Fretes", _money(fretes)],
        ["Saúde / sanidade", _money(custo_saude)],
        ["Despesas fixas operacionais (pro rata)", _money(custo_despesas_op)],
        ["Custo total operacional", _money(custo_total)],
    ], colWidths=[12 * cm, 5 * cm])
    custos_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.HexColor("#2D6A4F")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elems.append(custos_table)
    elems.append(Spacer(1, 0.4 * cm))

    # Resultado
    elems.append(Paragraph("Resultado", st["h2"]))
    res_color = "#1B4332" if lucro_liquido >= 0 else "#dc2626"
    resultado_table = Table([
        ["Lucro bruto (receita líquida – custos op.)", _money(lucro_bruto)],
        ["(–) Impostos (despesas fixas)", _money(custo_impostos)],
        ["Lucro líquido", _money(lucro_liquido)],
    ], colWidths=[12 * cm, 5 * cm])
    resultado_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEABOVE", (0, -1), (-1, -1), 1.0, colors.HexColor(res_color)),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 12),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor(res_color)),
    ]))
    elems.append(resultado_table)
    elems.append(Spacer(1, 0.6 * cm))

    # Detalhamento de movimentacoes
    elems.append(Paragraph(f"Movimentações do período ({len(movs)})", st["h2"]))
    if movs:
        rows = [["Data", "Tipo", "Brinco", "Valor", "Frete", "Desc.", "Observações"]]
        for m in movs:
            rows.append([
                m.data.strftime("%d/%m/%Y"),
                (m.tipo.value if hasattr(m.tipo, "value") else str(m.tipo)).title(),
                m.animal.brinco if m.animal else str(m.animal_id),
                _money(m.valor or 0),
                _money(m.frete or 0) if m.frete else "—",
                _money(m.desconto or 0) if m.desconto else "—",
                (m.observacoes or "")[:40],
            ])
        movs_table = Table(rows, repeatRows=1, colWidths=[2 * cm, 2 * cm, 2 * cm, 2.5 * cm, 2 * cm, 2 * cm, 5 * cm])
        movs_table.setStyle(_table_style())
        elems.append(movs_table)
    else:
        elems.append(Paragraph("Nenhuma movimentação no período.", st["muted"]))

    doc.build(elems)
    return _pdf_response(buf, f"resumo_contabil_{data_inicio}_{data_fim}.pdf")


@router.post("/animais/importar")
async def importar_animais(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(check_assinatura_ativa),
):
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .csv")

    contents = await file.read()
    try:
        text = contents.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = contents.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    criados = 0
    erros: list[str] = []
    brincos_no_csv: set[str] = set()  # detecta duplicatas dentro do proprio arquivo

    for i, row in enumerate(reader, start=2):
        brinco = (row.get("Brinco") or row.get("brinco") or "").strip()
        if not brinco:
            erros.append(f"Linha {i}: brinco ausente")
            continue

        if brinco in brincos_no_csv:
            erros.append(f"Linha {i}: brinco '{brinco}' duplicado no arquivo")
            continue

        existe = db.query(Animal).filter(
            Animal.user_id == current_user.id, Animal.brinco == brinco
        ).first()
        if existe:
            erros.append(f"Linha {i}: brinco '{brinco}' já existe")
            continue

        brincos_no_csv.add(brinco)

        sexo_raw = (row.get("Sexo") or row.get("sexo") or "macho").strip().lower()
        sexo = SexoEnum.femea if sexo_raw in ("femea", "fêmea", "f") else SexoEnum.macho

        animal = Animal(
            user_id=current_user.id,
            brinco=brinco,
            nome=(row.get("Nome") or row.get("nome") or "").strip() or None,
            raca=(row.get("Raça") or row.get("Raca") or row.get("raca") or "").strip() or None,
            sexo=sexo,
            status=StatusEnum.ativo,
            origem=(row.get("Origem") or row.get("origem") or "").strip() or None,
            observacoes=(row.get("Observações") or row.get("observacoes") or "").strip() or None,
        )

        peso_raw = (row.get("Peso Entrada (kg)") or row.get("peso_entrada") or "").strip()
        if peso_raw:
            try:
                p = float(peso_raw.replace(",", "."))
                if p > 0:
                    animal.peso_entrada = p
            except ValueError:
                pass

        db.add(animal)
        criados += 1

    db.commit()
    return {"importados": criados, "erros": erros}


# ── Livro Caixa (base para o LCDPR / IRPF Rural) ─────────────────────────────

CATEGORIA_DESP_LABEL = {
    "mao_de_obra": "Mão de obra",
    "manutencao": "Manutenção",
    "energia": "Energia elétrica",
    "arrendamento": "Arrendamento",
    "impostos": "Impostos",
    "sal_mineral": "Sal mineral",
    "suplemento": "Suplemento",
    "vermifugo": "Vermífugo",
    "combustivel": "Combustível",
    "outros": "Outros",
}


def _overlap(rec_inicio: date, rec_fim: Optional[date], per_inicio: date, per_fim: date) -> tuple[date, date] | None:
    """Retorna o intervalo (inicio, fim) de sobreposicao, ou None se nao houver."""
    inicio = max(rec_inicio, per_inicio)
    fim = min(rec_fim or per_fim, per_fim)
    return (inicio, fim) if fim >= inicio else None


@router.get("/livro-caixa.xlsx")
def exportar_livro_caixa(
    ano: int = Query(..., ge=2000, le=2100, description="Ano-calendário"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Livro Caixa do produtor rural para o ano-calendário informado.
    Agrega movimentações financeiras (compras/vendas), despesas fixas rateadas,
    custos de saúde e custos nutricionais rateados. Base para o LCDPR/IRPF Rural.
    """
    uid = current_user.id
    per_inicio = date(ano, 1, 1)
    per_fim = date(ano, 12, 31)

    linhas: list[dict] = []  # cada linha: data, tipo, categoria, descricao, referencia, valor

    # 1) Movimentacoes (compra = despesa, venda = receita)
    movs = (
        db.query(Movimentacao).join(Animal)
        .filter(
            Animal.user_id == uid,
            Movimentacao.data >= per_inicio,
            Movimentacao.data <= per_fim,
        )
        .order_by(Movimentacao.data)
        .all()
    )
    for m in movs:
        brinco = m.animal.brinco if m.animal else f"#{m.animal_id}"
        if m.tipo == TipoMovEnum.venda:
            valor = (m.valor or 0) - (m.desconto or 0)
            if valor > 0:
                linhas.append({
                    "data": m.data, "tipo": "Receita", "categoria": "Venda de animal",
                    "descricao": f"Venda — {m.destino or brinco}".strip(),
                    "referencia": brinco, "valor": valor,
                })
        elif m.tipo == TipoMovEnum.compra:
            valor_compra = (m.valor or 0) + (m.frete or 0)
            if valor_compra > 0:
                linhas.append({
                    "data": m.data, "tipo": "Despesa", "categoria": "Compra de animal",
                    "descricao": f"Compra — {m.origem or brinco}".strip(),
                    "referencia": brinco, "valor": valor_compra,
                })

    # 2) Custos de saude (com valor)
    saudes = (
        db.query(Saude).join(Animal)
        .filter(
            Animal.user_id == uid,
            Saude.custo != None,  # noqa: E711
            Saude.custo > 0,
            Saude.data >= per_inicio,
            Saude.data <= per_fim,
        )
        .order_by(Saude.data)
        .all()
    )
    for s in saudes:
        brinco = s.animal.brinco if s.animal else f"#{s.animal_id}"
        linhas.append({
            "data": s.data, "tipo": "Despesa", "categoria": "Saúde/Sanidade",
            "descricao": s.descricao or (s.tipo.value if hasattr(s.tipo, "value") else str(s.tipo)),
            "referencia": brinco, "valor": s.custo,
        })

    # 3) Despesas fixas (rateio mensal por dias sobrepostos)
    despesas = db.query(DespesaFixa).filter(DespesaFixa.user_id == uid).all()
    for d in despesas:
        ovl = _overlap(d.data_inicio, d.data_fim, per_inicio, per_fim)
        if not ovl:
            continue
        ini, fim = ovl
        # Gera 1 linha por mes de sobreposicao com o valor pro-rata do mes
        cur = date(ini.year, ini.month, 1)
        while cur <= fim:
            ultimo_dia = monthrange(cur.year, cur.month)[1]
            fim_mes = date(cur.year, cur.month, ultimo_dia)
            mes_ini = max(cur, ini)
            mes_fim = min(fim_mes, fim)
            dias_no_mes = (mes_fim - mes_ini).days + 1
            valor_mes = round((d.valor_mensal or 0) * dias_no_mes / ultimo_dia, 2)
            if valor_mes > 0:
                cat_label = CATEGORIA_DESP_LABEL.get(
                    d.categoria.value if hasattr(d.categoria, "value") else str(d.categoria),
                    "Outros",
                )
                linhas.append({
                    "data": fim_mes if fim_mes <= per_fim else per_fim,
                    "tipo": "Despesa", "categoria": cat_label,
                    "descricao": d.descricao,
                    "referencia": f"{mes_ini.strftime('%d/%m')} a {mes_fim.strftime('%d/%m')}",
                    "valor": valor_mes,
                })
            # avanca pro proximo mes
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)

    # 4) Custos nutricionais (rateio: preco_kg * consumo_kg_dia * dias * cabecas)
    custos_nutri = db.query(CustoNutricional).filter(CustoNutricional.user_id == uid).all()
    for c in custos_nutri:
        ovl = _overlap(c.data_inicio, c.data_fim, per_inicio, per_fim)
        if not ovl:
            continue
        ini, fim = ovl
        # Numero de cabecas: do lote se especifico, senao rebanho ativo total
        if c.lote_id:
            n_cabecas = db.query(sqlfunc.count(Animal.id)).filter(
                Animal.lote_id == c.lote_id,
                Animal.user_id == uid,
                Animal.deletado_em == None,  # noqa: E711
            ).scalar() or 0
            lote = db.query(Lote).filter(Lote.id == c.lote_id, Lote.user_id == uid).first()
            ref = f"Lote {lote.nome}" if lote else "Lote"
        else:
            n_cabecas = db.query(sqlfunc.count(Animal.id)).filter(
                Animal.user_id == uid,
                Animal.deletado_em == None,  # noqa: E711
                Animal.status == StatusEnum.ativo,
            ).scalar() or 0
            ref = "Rebanho"
        dias = (fim - ini).days + 1
        valor = round((c.preco_kg or 0) * (c.consumo_kg_dia or 0) * dias * n_cabecas, 2)
        if valor > 0:
            linhas.append({
                "data": fim, "tipo": "Despesa", "categoria": "Nutrição",
                "descricao": f"{c.produto} — {ini.strftime('%d/%m')} a {fim.strftime('%d/%m')}",
                "referencia": ref, "valor": valor,
            })

    # Ordena cronologicamente
    linhas.sort(key=lambda x: (x["data"], 0 if x["tipo"] == "Receita" else 1))

    # Calcula saldo acumulado
    saldo = 0.0
    for l in linhas:
        if l["tipo"] == "Receita":
            saldo += l["valor"]
        else:
            saldo -= l["valor"]
        l["saldo"] = round(saldo, 2)

    # ── Monta o XLSX ─────────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Livro Caixa cronologico
    ws = wb.active
    ws.title = "Livro Caixa"
    ws.append(["Data", "Tipo", "Categoria", "Descrição", "Referência", "Valor (R$)", "Saldo (R$)"])
    for l in linhas:
        ws.append([
            l["data"].isoformat(),
            l["tipo"],
            l["categoria"],
            l["descricao"],
            l["referencia"],
            l["valor"],
            l["saldo"],
        ])
    _format_header(ws)
    _autosize(ws)

    # Sheet 2: Resumo Mensal
    ws2 = wb.create_sheet("Resumo Mensal")
    ws2.append(["Mês", "Receitas (R$)", "Despesas (R$)", "Saldo do mês (R$)", "Saldo acumulado (R$)"])
    meses_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    saldo_acum = 0.0
    for mes in range(1, 13):
        rec = sum(l["valor"] for l in linhas if l["data"].month == mes and l["tipo"] == "Receita")
        desp = sum(l["valor"] for l in linhas if l["data"].month == mes and l["tipo"] == "Despesa")
        saldo_mes = rec - desp
        saldo_acum += saldo_mes
        ws2.append([
            f"{meses_nomes[mes - 1]}/{ano}",
            round(rec, 2), round(desp, 2), round(saldo_mes, 2), round(saldo_acum, 2),
        ])
    _format_header(ws2)
    _autosize(ws2)

    # Sheet 3: Resumo Anual + aviso
    ws3 = wb.create_sheet("Resumo Anual")
    total_rec = sum(l["valor"] for l in linhas if l["tipo"] == "Receita")
    total_desp = sum(l["valor"] for l in linhas if l["tipo"] == "Despesa")
    resultado = total_rec - total_desp
    ws3.append(["", ""])
    ws3.append(["Ano-calendário", ano])
    ws3.append(["Produtor", current_user.nome])
    ws3.append(["Fazenda", current_user.fazenda_nome or ""])
    ws3.append(["", ""])
    ws3.append(["Receita bruta total", round(total_rec, 2)])
    ws3.append(["Despesa total", round(total_desp, 2)])
    ws3.append(["Resultado da atividade rural", round(resultado, 2)])
    ws3.append(["", ""])
    ws3.append(["Total de lançamentos", len(linhas)])
    ws3.append(["", ""])
    ws3.append(["AVISO", "Este arquivo é um Livro Caixa gerencial baseado nos dados registrados no BovIA."])
    ws3.append(["", "Para gerar o LCDPR oficial (TXT layout 0002) ou entregar no IRPF Rural,"])
    ws3.append(["", "revise com seu contador. Custos nutricionais são rateios estimados."])
    _autosize(ws3)

    return _xlsx_response(wb, f"livro_caixa_{ano}.xlsx")
